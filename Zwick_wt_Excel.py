import pandas as pd
import os
from openpyxl import load_workbook
import numpy as np

# Set up paths to master folder and Excel file
master_folder_path = "D:/Nextcloud/Zugdehnung/Ergebnis_CSV"
excel_file_path = "D:/Nextcloud/Zugdehnung/Ergebnis_CSV/Zwickdaten.xlsx"

# Loop through subfolders in master folder
for folder_name in os.listdir(master_folder_path):
    folder_path = os.path.join(master_folder_path, folder_name)
    if os.path.isdir(folder_path):
        print(folder_path)
        # Create sheet name from folder name
        sheet_name = folder_name
        data_frames = []
        new_df = []
        stats_frames = []
        stats_res = []
        # Check if sheet with matching name already exists in Excel file
        existing_sheets = pd.read_excel(excel_file_path, sheet_name=None)
        if sheet_name in existing_sheets:

            # Loop through CSV files in folder
            for filename in os.listdir(folder_path):
                if filename.endswith(".csv") and filename.endswith("table.csv"):
                    # Read CSV file into a Pandas data frame
                    csv_path = os.path.join(folder_path, filename)
                    df = pd.read_csv(csv_path)
                    print('if 1')
                    # Add data frame to list for concatenation
                    data_frames.append(df)
            for filename in os.listdir(folder_path):
                if filename.endswith(".csv") and filename.endswith("stat.csv"):
                    # Read CSV file into a Pandas data frame
                    csv_path = os.path.join(folder_path, filename)
                    stats_df = pd.read_csv(csv_path)
                    print('if 2')

                    # Generate an array of indices to split the data frame
                    indices = np.array_split(np.arange(len(stats_df)), 10)

                    dfs = {}
                    for i, idx in enumerate(indices):
                        dfs[f'df_{i + 1}'] = stats_df.iloc[idx].reset_index(drop=True)

                    # Concatenate df_2, df_4, and df_8 into a new data frame

                    stat_res = pd.concat(
                        [dfs['df_10'], dfs['df_2'], dfs['df_8'], dfs['df_9'], dfs['df_1'], dfs['df_3'], dfs['df_4'],
                         dfs['df_5'], dfs['df_6'], dfs['df_7']], axis=1)

                    # Delete the third and fifth column of new_df
                    stat_res = stat_res.drop(stat_res.columns[[2, 4]], axis=1)
                    stat_res.insert(0, 'Stats', ['Average', 'Deviation', 'Variance'])
                    print(stat_res)
                    print('Replaced')
        else:
            # Sheet doesn't exist yet, create new sheet and write data to it

            # Loop through CSV files in folder
            for filename in os.listdir(folder_path):
                if filename.endswith(".csv") and filename.endswith("table.csv"):
                    print('else 1')
                    # Read CSV file into a Pandas data frame
                    csv_path = os.path.join(folder_path, filename)
                    df = pd.read_csv(csv_path)

                    # Add data frame to list for concatenation
                    data_frames.append(df)
            for filename in os.listdir(folder_path):
                if filename.endswith(".csv") and filename.endswith("stat.csv"):
                    # Read CSV file into a Pandas data frame
                    csv_path = os.path.join(folder_path, filename)
                    stats_df = pd.read_csv(csv_path)
                    print('else 2')
                    # Add data frame to list for concatenation
                    stats_frames.append(stats_df)

        # Concatenate all data frames into a new data frame for the new sheet
        new_df = pd.concat(data_frames, ignore_index=True)
        #st_df = pd.concat(stats_frames, ignore_index=True)

        #print(st_df)
        print('Written')

        # Write data frame to new sheet
        with pd.ExcelWriter(excel_file_path, engine="openpyxl", mode="a", if_sheet_exists='replace') as writer:
            writer.book = load_workbook(excel_file_path)
            new_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)



        with pd.ExcelWriter(excel_file_path, engine="openpyxl", mode="a", if_sheet_exists='overlay') as writer:
            writer.book = load_workbook(excel_file_path)
            stat_res.to_excel(writer, sheet_name=sheet_name, index=False, startrow=8, startcol=0)
            print('writing complete')

    print(folder_path)
print('Ende')