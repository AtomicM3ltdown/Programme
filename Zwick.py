import pandas as pd
import os
from openpyxl import load_workbook
import numpy as np

def replace_comma(s):
    return s.replace(',', '.')

def split_alpha_numeric(s):
    alpha = ''
    numeric = ''
    inside_braces = False
    for char in s:
        if char == '{':
            inside_braces = True
            alpha += char
        elif char == '}':
            inside_braces = False
            alpha += char
        elif char.isalpha() or inside_braces:
            alpha += char
        elif char.isnumeric() or (char == '.') or not inside_braces and not (char == ':'):
            if char != ' ':
                numeric += char
        elif char == ':':
            alpha += char
    if numeric == '':
        numeric = '0'
    return alpha, float(numeric)

# Get the list of files in the folder
folder_path = 'D:/Nextcloud/Zugdehnung/Ergebnis_CSV'
file_list = os.listdir(folder_path)

for filename in file_list:
    file_path = os.path.join(folder_path, filename)
    file_basename = os.path.splitext(filename)[0]
    dest_path = os.path.join(folder_path, file_basename)

    if filename.endswith(".TXT"):
        df1 = pd.DataFrame(columns = ['Stat', 'Value'])
        # file_path = "F:/Zugdehnung/JH_JCR1311.TXT"
        print(file_path)
        with open(file_path, 'r') as file:
            contents = file.read()
            contents = replace_comma(contents)

            # Move the file pointer to the beginning of the file
            file.seek(0)

            # Overwrite the file with the new content
            with open(file_path, 'w') as file:
                file.write(contents)

            # print(contents)
        # Load the CSV file into a DataFrame
        #df = pd.read_csv(file_path, on_bad_lines='skip', sep='\t')


        # Split the DataFrame into two based on the number of rows
        df1 = pd.read_csv(file_path, on_bad_lines='skip', sep='\t', encoding='unicode_escape', names=['Stat'],
                         engine='python').iloc[:31, :]

        df1[['Stat', 'Value']] = df1['Stat'].apply(lambda x: pd.Series(split_alpha_numeric(x)))

        df2 = pd.read_csv(file_path,
                          skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22,
                                    23,
                                    24, 25, 26, 27, 28, 29, 30],
                          sep='\t', on_bad_lines='skip',encoding= 'unicode_escape',
                          names=['\u03C3' + ' (MPa)',
                                 '\u03B5 Fmax' + ' (%)',
                                 'F (N)', '\u03C3R' + ' (MPa)','\u03B5' + ' (%)',
                                 '\u03B5 (50) (%)', '\u03B5 (100) (%)',
                                 '\u03B5 (200) (%)',
                                 '\u03B5 (300) (%)', '\u03B5 (500) (%)',
                                 'Lc (mm)', 'L0 (mm)', 'a (mm)', 'b (mm)',
                                 'Area (A\u2080) (mm\u00b2)'])
        # Add a sequential number to the 'Specimen number' column

        df2.insert(0, 'Specimen number', range(1, len(df2) + 1))


        #   Print out the two DataFrames to check that they look right
        print(df1)
        print(df2)
        # Create the destination folder if it doesn't exist
        if not os.path.exists(dest_path):
            os.makedirs(dest_path)

        df1.to_csv(os.path.join(dest_path, f"{file_basename}_stat.csv"), index=False, doublequote=False)
        df2.to_csv(os.path.join(dest_path, f"{file_basename}_table.csv"), index=False, doublequote=False)
        #df1 = None
        df2 = None

# Set up paths to master folder and Excel file
master_folder_path = "D:/Nextcloud/Zugdehnung/Ergebnis_CSV"
excel_file_path = "D:/Nextcloud/Zugdehnung/Ergebnis_CSV/Zwickdaten.xlsx"

# Loop through subfolders in master folder
for folder_name in os.listdir(master_folder_path):
    folder_path = os.path.join(master_folder_path, folder_name)
    if os.path.isdir(folder_path):
        print('x')
        # Create sheet name from folder name
        sheet_name = folder_name
        data_frames = []
        new_df = []
        stats_frames = []
        stats_res = []
        # Check if sheet with matching name already exists in Excel file
        existing_sheets = pd.read_excel(excel_file_path, sheet_name=None)
        if sheet_name in existing_sheets:

            # Loop through CSV files in folder
            for filename in os.listdir(folder_path):
                if filename.endswith(".csv") and filename.endswith("table.csv"):
                    # Read CSV file into a Pandas data frame
                    csv_path = os.path.join(folder_path, filename)
                    df = pd.read_csv(csv_path)
                    print('if 1')
                    # Add data frame to list for concatenation
                    data_frames.append(df)
            for filename in os.listdir(folder_path):
                if filename.endswith(".csv") and filename.endswith("stat.csv"):
                    # Read CSV file into a Pandas data frame
                    csv_path = os.path.join(folder_path, filename)
                    stats_df = pd.read_csv(csv_path)
                    print('if 2')

                    # Generate an array of indices to split the data frame
                    indices = np.array_split(np.arange(len(stats_df)), 10)

                    dfs = {}
                    for i, idx in enumerate(indices):
                        dfs[f'df_{i + 1}'] = stats_df.iloc[idx].reset_index(drop=True)

                    # Concatenate df_2, df_4, and df_8 into a new data frame

                    stat_res = pd.concat(
                        [dfs['df_10'], dfs['df_2'], dfs['df_8'], dfs['df_9'], dfs['df_1'], dfs['df_3'], dfs['df_4'],
                         dfs['df_5'], dfs['df_6'], dfs['df_7']], axis=1)

                    # Delete the third and fifth column of new_df
                    stat_res = stat_res.drop(stat_res.columns[[2, 4]], axis=1)
                    stat_res.insert(0, 'Stats', ['Average', 'Deviation', 'Variance'])
                    print(stat_res)
                    print('Replaced')
        else:
            # Sheet doesn't exist yet, create new sheet and write data to it

            # Loop through CSV files in folder
            for filename in os.listdir(folder_path):
                if filename.endswith(".csv") and filename.endswith("table.csv"):
                    print('else 1')
                    # Read CSV file into a Pandas data frame
                    csv_path = os.path.join(folder_path, filename)
                    df = pd.read_csv(csv_path)

                    # Add data frame to list for concatenation
                    data_frames.append(df)
            for filename in os.listdir(folder_path):
                if filename.endswith(".csv") and filename.endswith("stat.csv"):
                    # Read CSV file into a Pandas data frame
                    csv_path = os.path.join(folder_path, filename)
                    stats_df = pd.read_csv(csv_path)
                    print('else 2')
                    # Add data frame to list for concatenation
                    stats_frames.append(stats_df)

        # Concatenate all data frames into a new data frame for the new sheet
        new_df = pd.concat(data_frames, ignore_index=True)
        #st_df = pd.concat(stats_frames, ignore_index=True)

        #print(st_df)
        print('Written')

        # Write data frame to new sheet
        with pd.ExcelWriter(excel_file_path, engine="openpyxl", mode="a", if_sheet_exists='replace') as writer:
            writer.book = load_workbook(excel_file_path)
            new_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)



        with pd.ExcelWriter(excel_file_path, engine="openpyxl", mode="a", if_sheet_exists='overlay') as writer:
            writer.book = load_workbook(excel_file_path)
            stat_res.to_excel(writer, sheet_name=sheet_name, index=False, startrow=8, startcol=0)
            print('writing complete')

    print(folder_path)
print('Ende')
